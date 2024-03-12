"""======================================================
Please enter a URL in the following format: https://apps.apple.com/{region}/app/{app_id}
EXAMPLE: https://apps.apple.com/us/app/typeai-ai-keyboard-writer/id6448661220

I mean just copy link from the browser 😉
=======================================================\
"""

from bs4 import BeautifulSoup
import requests
import json
import re

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

import time
import random
import unicodedata
from datetime import datetime
from dataclasses import dataclass, fields


@dataclass
class CountryAppInfo():
    country: str
    total_ratings: str
    average_rating: str
    category_rank: str
    link: str


@dataclass
class AppReview():
    country: str
    date: str
    username: str
    rating: str
    title: str
    content: str
    link: str


class AppStoreParser:

    def __init__(self, domain: str, link_app_info: str, app_name: str):
        self.domain = domain
        self.link_app_info = link_app_info
        self.app_name = app_name

        # test dict
        # self.country_code_dict = {'': "United States", 'fr': 'France', 'de': 'Germany'}

        '''
        Down below I use the output of daily updated script that parses all 676 available ISO 3166-1 alpha-2 country
        codes and checks if it returns 200 OK status on iTunes.
        Link to repo: https://github.com/jcoester/iTunes-country-codes/tree/main

        P.S. I've tried parsing from official Apple website "Countries and Regions" page. 
        It works (code at the end of this file), but I deprecated it, because of problems like
        "uk" is for the main apple website, but for the app store website it's "gb" and etc.
        '''
        response = requests.get(
            'https://github.com/jcoester/iTunes-country-codes/blob/main/itunes_country_codes.json?raw=1')
        try:
            self.country_code_dict = response.json()
        except json.JSONDecodeError:  # just in case
            print("\n❗️ Country-code parsing failed. Changing to local country codes array (updated on 12.01.2024).\n")
            with open('itunes_country_codes.json', 'r') as file:
                self.country_code_dict = json.load(file)

    @staticmethod
    def get_soup_from_link(link: str):

        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                          'Chrome/58.0.3029.110 Safari/537.3',
        }

        app_store_response = requests.get(link, headers=headers)

        return BeautifulSoup(app_store_response.content, 'html.parser')

    def parse_app_info(self):
        print("======================================================="
              "\n❗️ Starting General App Info Parsing")
        self.common_parse(self.get_country_app_info)
        print("✅ General App Info Parsing has successfully ended."
              "\n=======================================================")

    def parse_app_reviews(self):
        print("======================================================="
              "\n❗️️ Starting App Reviews Parsing")
        self.common_parse(self.get_country_app_reviews)
        print("✅ App Reviews Parsing has successfully ended."
              "\n=======================================================")

    def common_parse(self, get_data_function):
        data_array = []
        for country_code, country in self.country_code_dict.items():
            data = get_data_function(country_code=country_code, country=country)
            # print(data)
            data_array.extend(data)

            delay = random.uniform(0.5, 1.0)
            time.sleep(delay)
        self.create_xlsx_file(array_of_data_objects=data_array)
        self.create_text_file(array_of_data_objects=data_array)
        return data_array

    # creating table column titles based on dataclass field names
    @staticmethod
    def generate_headers(class_object):

        current_date = datetime.now().strftime('%d.%m.%Y')

        titles_to_be_edited = {
            'category_rank': f'№ in Category on {current_date}',
            'content': f'Review Text ({current_date})'
        }

        table_headers = []
        for field in fields(class_object.__class__):
            title = snake_case_to_title(field.name)
            table_headers.append(titles_to_be_edited.get(field.name, title))

        return table_headers

    # universal function: array_of_data_objects type can be either 'CountryAppInfo' or 'AppReview'
    def create_xlsx_file(self, array_of_data_objects):

        if not array_of_data_objects:
            print("Looks like there's no data 🤔"
                  "\nTry another app!")

        else:
            is_countryappinfo_type = isinstance(array_of_data_objects[0],
                                              CountryAppInfo) # if false then it's AppReview type
            wb = Workbook()
            ws = wb.active

            table_headers = self.generate_headers(array_of_data_objects[0])

            # we don't include 'Link' header because there'll be clickable link in Country header
            ws.append(table_headers[:-1])

            for data_object in array_of_data_objects:
                row = list(vars(data_object).values())  # starting from Python3.7 dictionaries are ordered datastructure
                ws.append(row[:-1])  # we don't  include link as separate column

                cell = ws.cell(row=ws.max_row, column=1)
                cell.hyperlink = data_object.link
                cell.font = Font(color="0000FF")

            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            ws.column_dimensions['A'].width = 17
            ws.column_dimensions['B'].width = 17
            ws.column_dimensions['C'].width = 17
            ws.column_dimensions['D'].width = 20

            if is_countryappinfo_type is False:
                ws.column_dimensions['E'].width = 30
                ws.column_dimensions['F'].width = 65

            file_purpose = 'info' if is_countryappinfo_type else 'reviews'

            filename = f"{self.app_name}_{file_purpose}_{datetime.now().strftime('%d-%m-%Y_%H-%M-%S')}.xlsx"
            wb.save(filename)

            print(f'📁 File "{filename}" has been successfully saved on disk')

    def create_text_file(self, array_of_data_objects):
        if not array_of_data_objects:
            print("Looks like there's no data 🤔\nTry another app!")
            return

        is_countryappinfo_type = isinstance(array_of_data_objects[0], CountryAppInfo)
        file_purpose = 'info' if is_countryappinfo_type else 'reviews'
        filename = f"{self.app_name}_{file_purpose}_{datetime.now().strftime('%d-%m-%Y_%H-%M-%S')}.txt"

        with open(filename, 'w') as file:
            for data_object in array_of_data_objects:
                for field in fields(data_object.__class__):
                    # Skip adding link to the text file content
                    if field.name != "link":
                        value = getattr(data_object, field.name)
                        file.write(f"{snake_case_to_title(field.name)}: '{value}'\n")
                file.write("\n")  # Extra newline for separating entries

        print(f'📁 File "{filename}" has been successfully saved on disk')

    def get_country_app_info(self, country_code: str, country: str):

        current_country_link = self.domain + country_code + self.link_app_info
        # print(current_country_link)

        soup = self.get_soup_from_link(current_country_link)

        print(f'Analyzing general app info in {country}')

        category_ranking_section = soup.find('a', href=lambda x: x and 'charts' in x)
        category_ranking = unicodedata.normalize("NFKC",
                                                 category_ranking_section.get_text().strip()) \
            if category_ranking_section else '-'

        rating_value = 'Not found'
        total_ratings = 'Not found'

        rating_section = soup.find('figure', class_='we-star-rating')
        if rating_section:
            figcaption = rating_section.find('figcaption', class_='we-rating-count star-rating__count')
            if figcaption:
                rating_text = figcaption.get_text()
                rating = unicodedata.normalize("NFKC", rating_text)
                rating_value, total_ratings = rating.split(' • ')

        return [CountryAppInfo(country=country, average_rating=rating_value, total_ratings=total_ratings,
                              category_rank=category_ranking, link=current_country_link)]

    def get_country_app_reviews(self, country_code: str, country: str):

        current_country_link = self.domain + country_code + self.link_app_info + '?see-all=reviews'
        # print(current_country_link)

        soup = self.get_soup_from_link(current_country_link)

        country_app_reviews_array: [AppReview] = []

        print(f'Searching for reviews in {country}')

        for review_div in soup.find_all("div", class_="we-customer-review lockup"):

            username_section = review_div.find("span", class_="we-customer-review__user")
            review_username = username_section.text.strip() if username_section else "No Username"

            date_section = review_div.find("time", class_="we-customer-review__date")
            review_date = date_section['datetime'] if date_section else "No Date"

            rating_section = review_div.find("figure", class_="we-star-rating")
            if rating_section:
                rating_text = rating_section['aria-label']
                match = re.search(r'(\d)[^0-9]+(\d)', rating_text)
                review_rating = match.group(1) if match else "No Rating"
            else:
                review_rating = "No Rating"

            title_section = review_div.find("h3", class_="we-customer-review__title")
            review_title = title_section.text.strip() if title_section else "No Title"

            content_section = review_div.find("blockquote", class_="we-customer-review__body")
            review_content = content_section.text.strip() if content_section else "No Content"

            country_app_reviews_array.append(
                AppReview(username=review_username, date=review_date, rating=review_rating, title=review_title,
                          content=review_content, country=country, link=current_country_link))

        return country_app_reviews_array


def validate_url(pattern: str):
    while True:
        url = input("Enter the App Store URL: ")
        matches = re.match(pattern=pattern, string=url)
        if matches:
            return matches.groups()
        print("\nThat's not a valid App Store URL 😓")


def get_user_action():
    while True:
        action = input("Just type the necessary digit and press 'Enter':"
                       "\n1 - Parse basic app info (Country, Number of Ratings, App Store Rating, # in Category)"
                       " for all countries"
                       "\n2 - Parse app reviews from the web version of the App Store for all countries"
                       " (Country, Date, Username, Rating, Title, Review Text)"
                       "\n3 - Both options"
                       "\n0 - Return to link entering step\n")
        if action in ["0", "1", "2", "3"]:
            return action
        print("Oh, that's not a correct digit (1, 2, 3, or 0) 😓"
              "\nLet's try again!")


def snake_case_to_title(string: str):
    return ' '.join(word.capitalize() for word in string.split('_'))


def execute_user_action(domain: str, link_app_info: str, app_name: str, action: str):
    if action == "0":
        return

    parser = AppStoreParser(domain=domain, link_app_info=link_app_info, app_name=app_name)

    if action == "1":
        parser.parse_app_info()
    elif action == "2":
        parser.parse_app_reviews()
    elif action == "3":
        parser.parse_app_info()
        parser.parse_app_reviews()


def main():
    pattern = r'(https://apps\.apple\.com/)\w+(/app/([\w\-]+)/.*)'
    while True:
        print(__doc__)
        domain, link_app_info, app_name = validate_url(pattern=pattern)
        action = get_user_action()
        execute_user_action(domain=domain, link_app_info=link_app_info, app_name=app_name, action=action)


if __name__ == '__main__':
    main()

'''
This method of parsing from official Apple website "Countries and Regions" page works, but deprecated
because of problems like 'uk' is for the main apple website, but for the app store website it's 'gb'
'''
# def fetch_app_store_countries():
#
#     url = 'https://www.apple.com/choose-country-region/'
#
#     response = requests.get(url, headers=headers)
#     soup = BeautifulSoup(response.content, 'html.parser')
#
#     a_tags = soup.find_all('a', {'property': 'schema:url'})
#
#     country_code_dict = {}
#     for tag in a_tags:
#
#         country_name = tag.find('span', {'property': 'schema:name'}).text
#         country_code = tag['href'].strip('/')
#
#         country_code_dict[country_code] = country_name
#
#     return country_code_dict

'''
using method (deprecated)
'''
# # country_code_dict = fetch_app_store_countries()
# if not country_code_dict: # just in case
#     print("Country-code parsing failed. Changing to local country codes array (updated on 12.01.2024).")
#     country_code_dict = {'bh': 'Bahrain', 'bh-ar': 'البحرين', 'bw': 'Botswana', 'cm': 'Cameroun',
#                          'cf': 'République Centrafricaine', 'ci': "Côte d'Ivoire", 'eg': 'Egypt',
#                          'eg-ar': 'مصر', 'gw': 'Guinea-Bissau', 'gn': 'Guinée', 'gq': 'Guinée Equatoriale',
#                          'in': 'India', 'il': 'Israel', 'jo': 'Jordan', 'jo-ar': 'الأردن', 'ke': 'Kenya',
#                          'kw': 'Kuwait', 'kw-ar': 'الكويت', 'mg': 'Madagascar', 'ml': 'Mali', 'ma': 'Maroc',
#                          'mu': 'Maurice', 'mz': 'Mozambique', 'ne': 'Niger', 'ng': 'Nigeria', 'om': 'Oman',
#                          'om-ar': 'عُمان', 'qa': 'Qatar', 'qa-ar': 'قطر', 'sa': 'Saudi Arabia',
#                          'sa-ar': 'المملكة العربية السعودية', 'sn': 'Sénégal', 'za': 'South Africa',
#                          'tn': 'Tunisie', 'ug': 'Uganda', 'ae': 'United Arab Emirates',
#                          'ae-ar': 'الإمارات العربية المتحدة', 'au': 'Australia',
#                          'https://www.apple.com.cn': '中国大陆', 'hk/en': 'Hong Kong', 'hk': '香港',
#                          'id': 'Indonesia', 'jp': '日本', 'kr': '대한민국', 'mo': '澳門', 'my': 'Malaysia',
#                          'nz': 'New Zealand', 'ph': 'Philippines', 'sg': 'Singapore', 'tw': '台灣', 'th': 'ไทย',
#                          'vn': 'Việt Nam', 'am': 'Armenia', 'az': 'Azerbaijan', 'by': 'Belarus',
#                          'benl': 'België', 'befr': 'Belgique', 'bg': 'България', 'cz': 'Česko', 'dk': 'Danmark',
#                          'de': 'Deutschland', 'ee': 'Eesti', 'es': 'España', 'fr': 'France', 'ge': 'Georgia',
#                          'gr': 'Ελλάδα', 'hr': 'Hrvatska', 'ie': 'Ireland', 'it': 'Italia', 'kz': 'Kazakhstan',
#                          'kg': 'Kyrgyzstan', 'lv': 'Latvija', 'li': 'Liechtenstein', 'lt': 'Lietuva',
#                          'lu': 'Luxembourg', 'hu': 'Magyarország', 'mt': 'Malta', 'md': 'Moldova',
#                          'me': 'Montenegro', 'nl': 'Nederland', 'mk': 'North Macedonia', 'no': 'Norge',
#                          'at': 'Österreich', 'pl': 'Polska', 'pt': 'Portugal', 'ro': 'România',
#                          'sk': 'Slovensko', 'si': 'Slovenia', 'chde': 'Schweiz', 'chfr': 'Suisse',
#                          'fi': 'Suomi', 'se': 'Sverige', 'tj': 'Tajikistan', 'tr': 'Türkiye',
#                          'tm': 'Turkmenistan', 'uk': 'United Kingdom', 'ua': 'Україна', 'uz': 'Uzbekistan',
#                          'lae': 'Puerto Rico (English)', 'la': 'Puerto Rico (Español)', 'br': 'Brasil',
#                          'cl': 'Chile', 'co': 'Colombia', 'mx': 'México', 'ca': 'Canada (English)',
#                          'ca/fr': 'Canada (Français)', '': 'United States'}
